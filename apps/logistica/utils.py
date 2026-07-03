import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

def exportar_kardex_excel(item, movimientos, saldo_inicial):
    """Genera un archivo Excel profesional con el historial del Kardex."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kardex de Inventario"

    # Título y Saldo Inicial
    ws.merge_cells('A1:G1')
    ws['A1'] = f"REPORTE KARDEX: {item.nombre.upper()}"
    ws['A1'].font = Font(bold=True, size=14, color="1e293b")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = "SALDO INICIAL ACUMULADO:"
    ws['B2'] = float(saldo_inicial)
    ws['B2'].font = Font(bold=True)

    # Cabeceras
    headers = ["Fecha", "Documento", "Motivo", "Entrada", "Salida", "Saldo", "Costo Unit."]
    ws.append([]) # Fila vacía
    ws.append(headers)

    header_fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Filas de datos
    for idx, mov in enumerate(movimientos, 5):
        ws.cell(row=idx, column=1, value=mov['fecha'].strftime('%d/%m/%Y %H:%M') if mov['fecha'] else "")
        ws.cell(row=idx, column=2, value=mov['documento'])
        ws.cell(row=idx, column=3, value=mov['motivo'])
        ws.cell(row=idx, column=4, value=float(mov['entrada']))
        ws.cell(row=idx, column=5, value=float(mov['salida']))
        ws.cell(row=idx, column=6, value=float(mov['saldo']))
        ws.cell(row=idx, column=7, value=float(mov['costo']))

    # Ajuste automático de columnas
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="kardex_{item.id}.xlsx"'
    wb.save(response)
    return response